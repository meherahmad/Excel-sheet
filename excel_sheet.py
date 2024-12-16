#PROJECT 1
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('transactions.xlsx')  # load our excel workbook
sheet = wb['Sheet1']  # access sheet 1
# cell = sheet['a1']  # access cell a1
# print(cell.value)   # prints the value in cell a1
# print(sheet.max_row) # prints the max number of rows

for row in range(2, sheet.max_row + 1):  # iterates from row 2 to the last row
    cell = sheet.cell(row, 3)  # for each row this access the cell in the 3rd column
    print(cell.value)  # print all the values from the 3rd column
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)  # accesses cell in the 4th column for all the rows
    corrected_price_cell.value = corrected_price  # assigns the corrected values to the 4th column

values = Reference(sheet,      # creates a reference object
                   min_row=2,
                   max_row=sheet.max_row,
                   min_col=4,
                   max_col=4)
                # this range will be used as the data source for the bar chart.
# creates a barchart in the spreadsheet
chart = BarChart()   # create an instance of the barchart class
chart.add_data(values)
sheet.add_chart(chart, 'F2')  # places the chart in the sheet starting at cell F2

wb.save('transactions2.xlsx')  # saves the modified data and the new chart into a new exel file.